import os
import uuid
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from typing import List, Dict, Optional
from data_intelligence_system.utils.logger import get_logger

logger = get_logger("ExcelReportGenerator")


class ExcelReportGenerator:
    def __init__(self, filename: str):
        """
        إنشاء تقرير Excel جديد.
        :param filename: مسار واسم ملف الإكسل الناتج (.xlsx)
        """
        self.filename = filename
        self.wb = openpyxl.Workbook()
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)

    def _set_column_widths(self, ws, widths: Dict[int, int]):
        for col_idx, width in widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    def _apply_border(self, cell):
        thin = Side(style='thin')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _apply_header_style(self, cell):
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6EFD")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_section_title_style(self, cell):
        cell.font = Font(bold=True, size=12, color="0D6EFD")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def _apply_table_header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6EFD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        self._apply_border(cell)

    def _apply_table_cell_style(self, cell):
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self._apply_border(cell)

    def _write_section(self, ws, start_row: int, section: Dict) -> int:
        row = start_row
        title_cell = ws.cell(row=row, column=1, value=section.get("title", ""))
        self._apply_section_title_style(title_cell)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 2

        paragraphs = section.get("paragraphs", [])
        for para in paragraphs:
            para_cell = ws.cell(row=row, column=1, value=para)
            para_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="right")
            ws.row_dimensions[row].height = 30
            row += 1
        if paragraphs:
            row += 1

        tables = section.get("tables", [])
        for table_idx, table in enumerate(tables):
            headers = table.get("headers", [])
            rows_data = table.get("rows", [])

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                self._apply_table_header_style(cell)
            row += 1

            for row_data in rows_data:
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    self._apply_table_cell_style(cell)
                row += 1

            if headers and rows_data:
                table_range = f"A{row - len(rows_data) -1}:{get_column_letter(len(headers))}{row - 1}"
                table_id = uuid.uuid4().hex[:8]
                tab = Table(displayName=f"Table_{start_row}_{table_idx}_{table_id}", ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)

            row += 2

        return row

    def generate(self, title: str, sections: List[Dict], author: Optional[str] = None):
        if not sections:
            logger.warning("⚠️ القائمة sections فارغة، لن يتم إنشاء تقرير.")
            return

        os.makedirs(os.path.dirname(self.filename), exist_ok=True)

        if author:
            self.wb.properties.creator = author
        self.wb.properties.title = title
        self.wb.properties.created = datetime.now()

        sheet_name = title[:31] if title else "Report"
        ws = self.wb.create_sheet(title=sheet_name)

        ws.merge_cells("A1:J1")
        header_cell = ws["A1"]
        header_cell.value = title
        self._apply_header_style(header_cell)
        ws.row_dimensions[1].height = 30

        self._set_column_widths(ws, {i: 20 for i in range(1, 11)})
        current_row = 3

        for section in sections:
            current_row = self._write_section(ws, current_row, section)

        try:
            self.wb.save(self.filename)
            logger.info(f"✅ تم إنشاء تقرير Excel وحفظه في: {self.filename}")
        except Exception as e:
            logger.error(f"❌ خطأ أثناء حفظ ملف الإكسل: {e}", exc_info=True)
